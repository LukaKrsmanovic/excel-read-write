import openpyxl
import pandas as pd
from indecies import grade_index, pass_index
from openpyxl.styles import PatternFill

LOAD_FILE_PATH = './documents/test.xlsx'
SAVE_FILE_PATH = './documents/test_modified.xlsx'

wb = openpyxl.load_workbook(LOAD_FILE_PATH)
ws = wb.active

fill = PatternFill(start_color="6bff7c", end_color="6bff7c", fill_type="solid")
for row in ws.iter_rows(min_row=2):
    grade = row[grade_index].value
    if grade >= 50:
        row[pass_index].value = 'PASS'
        for cell in row:
            cell.fill = fill
    else:
        row[pass_index].value = 'FAIL'
wb.save(SAVE_FILE_PATH)
